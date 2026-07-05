import json
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

EXPORT_K = 6  # broader retrieval than normal chat (k=2), since we need full statement coverage

FIELDS = [
    "fiscal_year",
    "total_revenue",
    "net_income",
    "gross_margin",
    "operating_income",
    "eps",
    "total_assets",
    "total_liabilities",
]

FIELD_LABELS = {
    "fiscal_year": "Fiscal Year",
    "total_revenue": "Total Revenue",
    "net_income": "Net Income / PAT",
    "gross_margin": "Gross Margin",
    "operating_income": "Operating Income",
    "eps": "EPS",
    "total_assets": "Total Assets",
    "total_liabilities": "Total Liabilities",
}

EXTRACTION_QUERY = (
    "total revenue net income profit after tax gross margin operating income "
    "earnings per share total assets total liabilities fiscal year"
)


def extract_financial_summary(llm, vectorstore, company_name: str) -> dict:
    """
    Pull a structured set of headline financial figures for one company directly
    from the shared vectorstore (filtered by company), independent of chat history.
    Returns values exactly as reported in the filing -- no unit/currency conversion,
    no guessing. Missing fields come back as "N/A".
    """
    docs = vectorstore.similarity_search(EXTRACTION_QUERY, k=EXPORT_K, filter={"company": company_name})
    context = "\n\n".join(d.page_content for d in docs)

    prompt = (
        "You are extracting financial figures from the following excerpt of a company's "
        "financial report. Respond with ONLY a valid JSON object, no markdown fences, no "
        "commentary, no explanation -- just the raw JSON.\n\n"
        f"Required keys: {FIELDS}\n\n"
        "Rules:\n"
        "- Report each value EXACTLY as it appears in the text, including currency symbol "
        "and unit (e.g. '$391,035 million' or '\u20b91,20,000 crore'). Do not convert units "
        "or currencies.\n"
        "- \"gross_margin\" means gross profit (total revenue minus cost of revenue/cost of goods "
        "sold), NOT the cost of revenue itself. If the filing states a gross margin percentage, "
        "include that too, e.g. '$195,201 million (46.9%)'.\n"
        "- If a figure is not present in the excerpt, use the string \"N/A\" for that key. "
        "Do not guess or estimate.\n\n"
        f"Excerpt:\n{context}"
    )

    result = {field: "N/A" for field in FIELDS}
    try:
        response = llm.invoke(prompt)
        text = response.content.strip()
        # Strip markdown code fences if the model added them anyway
        if text.startswith("```"):
            text = text.strip("`")
            if text.lower().startswith("json"):
                text = text[4:]
        parsed = json.loads(text.strip())
        for field in FIELDS:
            value = parsed.get(field)
            if value and str(value).strip():
                result[field] = str(value).strip()
    except Exception:
        # Leave everything as "N/A" rather than surface a malformed/partial result
        pass

    return result


def build_excel_bytes(summaries: list[dict]) -> bytes:
    """Build an .xlsx file (in memory) from a list of dicts produced by extract_financial_summary."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Financial Summary"

    headers = ["Company"] + [FIELD_LABELS[f] for f in FIELDS] + ["Source Report"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in summaries:
        ws.append(
            [row["company"]] + [row.get(f, "N/A") for f in FIELDS] + [row["source_report"]]
        )

    for i, _ in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = 20

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
